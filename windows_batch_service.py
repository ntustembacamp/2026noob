import csv
import json
import os
import sys
import shutil
import ctypes
import subprocess
import contextlib
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel, Field


BASE_DIR = Path(__file__).resolve().parent
SERVICE_DIR = BASE_DIR / "service"
if str(SERVICE_DIR) not in sys.path:
    sys.path.insert(0, str(SERVICE_DIR))

from activity_workflows import (
    normalize_activity_photo_files,
    normalize_activity_photo_folder,
    import_activity_photos_from_normalized_folder,
    retry_failed_activity_recognition,
)
from log_paths import (
    ACTIVITY_PHOTO_IMPORT_LOG_PATH,
    FEATURE_BUILD_LOG_PATH,
    LEGACY_LOG_ROOT,
    LEGACY_MIGRATED_LOG_ROOT,
    NOOB_LOG_ROOT,
    WINDOWS_BATCH_NORMALIZE_LOG_PATH,
    ensure_noob_log_root,
)

BATCH_UPLOAD_ROOT = BASE_DIR / "database" / "tmp_batch_uploads_windows"
BATCH_OUTPUT_ROOT = BASE_DIR / "database" / "tmp_batch_outputs_windows"
LOG_ROOT = NOOB_LOG_ROOT


def _proxy_error_message(exc: Exception, action: str) -> str:
    if isinstance(exc, urllib.error.HTTPError):
        detail = ""
        with contextlib.suppress(Exception):
            raw = exc.read().decode("utf-8", errors="replace")
            if raw:
                parsed = json.loads(raw)
                detail = str(parsed.get("detail", "")).strip()
        if exc.code == 422:
            if "schedule_id" in detail or "int_parsing" in detail:
                return f"{action}失敗：活動行程欄位格式錯誤（schedule_id），請重新選擇活動行程後再試。"
            return f"{action}失敗：欄位格式錯誤（HTTP 422）"
        return f"{action}失敗：HTTP {exc.code}{(' - ' + detail) if detail else ''}"
    if isinstance(exc, urllib.error.URLError):
        reason = getattr(exc, "reason", None)
        return f"{action}失敗：無法連線到 8000 服務（{reason}）"
    message = str(exc).strip()
    if not message:
        message = repr(exc)
    return f"{action}失敗：{message}"


app = FastAPI(title="Noob Windows Batch Service", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8000",
        "http://127.0.0.1:8000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class BatchNormalizePayload(BaseModel):
    source_dir: str = Field(..., min_length=1)
    destination_dir: str = Field(default="")
    excel_path: str = Field(..., min_length=1)
    sheet_name: str = Field(default="")
    original_filename_column: str = Field(..., min_length=1)
    filename_fields: list[str] = Field(..., min_length=1)
    delimiter: str = Field(default="_")
    extension_override: str = Field(default="")


class OpenLogPayload(BaseModel):
    log_name: str = Field(..., min_length=1)


class OpenFolderPayload(BaseModel):
    folder_path: str = Field(..., min_length=1)


def ensure_batch_roots():
    BATCH_UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    BATCH_OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    ensure_noob_log_root()


def write_log_file(log_path: Path, lines: list[str]):
    ensure_noob_log_root()
    with log_path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write("\n".join(lines or []))


def get_recent_legacy_log_files(seconds: int = 60) -> list[Path]:
    if not LEGACY_LOG_ROOT.exists():
        return []
    now = datetime.now().timestamp()
    recent_files = []
    for path in LEGACY_LOG_ROOT.glob("*.log"):
        with contextlib.suppress(Exception):
            if now - path.stat().st_mtime <= seconds:
                recent_files.append(path)
    return sorted(recent_files, key=lambda p: p.name.lower())


def create_batch_job_dir(root: Path, prefix: str):
    ensure_batch_roots()
    job_dir = root / f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}"
    job_dir.mkdir(parents=True, exist_ok=True)
    return job_dir


def save_upload_file(upload: UploadFile, target_path: Path):
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with target_path.open("wb") as buffer:
        shutil.copyfileobj(upload.file, buffer)


def safe_filename_part(value):
    text = str(value or "").strip()
    for char in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
        text = text.replace(char, "_")
    return text


def build_normalized_filename(row, filename_fields, delimiter, original_ext, extension_override):
    filename_parts = [safe_filename_part(row.get(field, "")) for field in filename_fields]
    filename_parts = [part for part in filename_parts if part]
    normalized_stem = delimiter.join(filename_parts)
    extension = extension_override.strip() if extension_override else original_ext
    if extension and not extension.startswith("."):
        extension = f".{extension}"
    return f"{normalized_stem}{extension}"


def build_unique_archive_path(target_dir: Path, original_name: str):
    candidate = target_dir / original_name
    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix
    counter = 1
    while True:
        candidate = target_dir / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def build_unique_file_path(target_dir: Path, original_name: str):
    candidate = target_dir / original_name
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    counter = 1
    while True:
        candidate = target_dir / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def preserve_file_times(source_file: Path, target_file: Path):
    source_stat = source_file.stat()
    os.utime(target_file, ns=(source_stat.st_atime_ns, source_stat.st_mtime_ns))

    if os.name != "nt":
        return

    FILE_WRITE_ATTRIBUTES = 0x0100
    OPEN_EXISTING = 3

    class FILETIME(ctypes.Structure):
        _fields_ = [("dwLowDateTime", ctypes.c_uint32), ("dwHighDateTime", ctypes.c_uint32)]

    def to_filetime(timestamp_ns: int) -> FILETIME:
        value = int(timestamp_ns // 100) + 116444736000000000
        return FILETIME(value & 0xFFFFFFFF, value >> 32)

    handle = ctypes.windll.kernel32.CreateFileW(
        str(target_file),
        FILE_WRITE_ATTRIBUTES,
        0,
        None,
        OPEN_EXISTING,
        0,
        None,
    )
    if handle in (0, ctypes.c_void_p(-1).value):
        return
    try:
        creation_time = to_filetime(source_stat.st_ctime_ns)
        ctypes.windll.kernel32.SetFileTime(handle, ctypes.byref(creation_time), None, None)
    finally:
        ctypes.windll.kernel32.CloseHandle(handle)


def list_tabular_sheets(file_path: str | Path):
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return ["CSV"]
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        return workbook.sheetnames
    raise ValueError("僅支援 Excel 或 CSV 檔案。")


def read_tabular_rows(file_path: str | Path, sheet_name: str = ""):
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            columns = reader.fieldnames or []
            rows = [{str(k): "" if v is None else str(v) for k, v in row.items()} for row in reader]
            return columns, rows

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    target_sheet = sheet_name or workbook.sheetnames[0]
    if target_sheet not in workbook.sheetnames:
        raise ValueError(f"找不到工作表：{target_sheet}")

    worksheet = workbook[target_sheet]
    values = list(worksheet.iter_rows(values_only=True))
    if not values:
        return [], []

    columns = ["" if value is None else str(value).strip() for value in values[0]]
    rows = []
    for raw_row in values[1:]:
        row = {}
        for index, column in enumerate(columns):
            if not column:
                continue
            cell_value = raw_row[index] if index < len(raw_row) else ""
            row[column] = "" if cell_value is None else str(cell_value).strip()
        rows.append(row)
    return columns, rows


def load_excel_columns(excel_path: str, sheet_name: str = ""):
    columns, _ = read_tabular_rows(excel_path, sheet_name=sheet_name)
    return [str(column) for column in columns if str(column).strip()]


def resolve_source_file(source_path: Path, original_name: str):
    direct_match = source_path / original_name
    if direct_match.is_file():
        return direct_match

    original_path = Path(original_name)
    target_stem = original_path.stem if original_path.suffix else original_name
    matches = sorted(
        candidate for candidate in source_path.iterdir() if candidate.is_file() and candidate.stem == target_stem
    )
    if matches:
        return matches[0]
    return None


def archive_processed_source_files(source_path: Path, processed: list[dict]):
    archive_dir = source_path / "_normalized_success"
    archive_dir.mkdir(parents=True, exist_ok=True)

    archived = []
    for item in processed:
        source_file = Path(item["source_file"])
        archive_path = build_unique_archive_path(archive_dir, source_file.name)
        shutil.copy2(source_file, archive_path)
        preserve_file_times(source_file, archive_path)
        if source_file.exists():
            source_file.unlink()
        archived.append(
            {
                "original_name": source_file.name,
                "archive_file": str(archive_path),
            }
        )

    return archive_dir, archived


def normalize_headshot_batch(payload: BatchNormalizePayload):
    source_path = Path(payload.source_dir)
    destination_path = Path(payload.destination_dir) if payload.destination_dir else create_batch_job_dir(
        BATCH_OUTPUT_ROOT, "normalized"
    )
    excel_file = Path(payload.excel_path)

    if not source_path.exists():
        source_path.mkdir(parents=True, exist_ok=True)
    if not source_path.is_dir():
        raise ValueError("來源圖檔路徑不是資料夾")
    if not excel_file.is_file():
        raise ValueError("Excel 檔案不存在")
    if not payload.filename_fields:
        raise ValueError("至少要選一個正規化檔名欄位")

    destination_path.mkdir(parents=True, exist_ok=True)
    columns, rows = read_tabular_rows(excel_file, sheet_name=payload.sheet_name)

    if payload.original_filename_column not in columns:
        raise ValueError("Excel 找不到大頭照原始檔名欄位")
    for field in payload.filename_fields:
        if field not in columns:
            raise ValueError(f"Excel 找不到欄位：{field}")

    processed = []
    missing = []
    duplicated_targets = []
    logs = [f"{datetime.now().isoformat(sep=' ', timespec='seconds')} 開始執行圖檔正規化"]
    logs.append(f"來源資料夾：{source_path}")
    logs.append(f"目的資料夾：{destination_path}")
    logs.append(f"Excel：{excel_file}")
    target_names_seen = set()

    for row in rows:
        original_name = str(row.get(payload.original_filename_column, "")).strip()
        if not original_name:
            continue

        source_file = resolve_source_file(source_path, original_name)
        if source_file is None:
            missing.append(original_name)
            logs.append(f"缺少來源檔：{original_name}")
            continue

        normalized_name = build_normalized_filename(
            row=row,
            filename_fields=payload.filename_fields,
            delimiter=payload.delimiter or "_",
            original_ext=source_file.suffix,
            extension_override=payload.extension_override,
        )
        target_file = destination_path / normalized_name

        if normalized_name in target_names_seen:
            duplicated_targets.append(normalized_name)
            logs.append(f"目標檔名重複，略過：{normalized_name}")
            continue
        target_names_seen.add(normalized_name)

        shutil.copy2(source_file, target_file)
        preserve_file_times(source_file, target_file)
        processed.append(
            {
                "original_name": original_name,
                "normalized_name": normalized_name,
                "source_file": str(source_file),
                "target_file": str(target_file),
            }
        )
        logs.append(f"已完成：{original_name} -> {normalized_name}")

    archive_dir, archived = archive_processed_source_files(source_path, processed)
    logs.append(f"完成：成功 {len(processed)} 筆，缺檔 {len(missing)} 筆，重複目標 {len(duplicated_targets)} 筆")
    logs.append(f"備份資料夾：{archive_dir}")
    write_log_file(WINDOWS_BATCH_NORMALIZE_LOG_PATH, logs)
    return {
        "destination_dir": str(destination_path),
        "destination_host_dir": str(destination_path),
        "archive_dir": str(archive_dir),
        "archive_host_dir": str(archive_dir),
        "archived_count": len(archived),
        "processed_count": len(processed),
        "missing_count": len(missing),
        "duplicate_target_count": len(duplicated_targets),
        "processed": processed[:100],
        "archived_files": archived[:100],
        "missing_files": missing[:100],
        "duplicate_targets": duplicated_targets[:100],
        "logs": logs[:500],
        "log_path": str(WINDOWS_BATCH_NORMALIZE_LOG_PATH),
    }


@app.get("/health")
async def health():
    recent_legacy = get_recent_legacy_log_files(60)
    return {
        "ok": True,
        "proxy_mode": "start_and_poll",
        "service_code_version": "2026-05-20-activity-import-proxy-v2",
        "log_root": str(NOOB_LOG_ROOT),
        "service_cwd": os.getcwd(),
        "legacy_log_root": str(LEGACY_LOG_ROOT),
        "legacy_log_recent_count": len(recent_legacy),
        "legacy_log_recent_files": [str(item) for item in recent_legacy[:20]],
    }


@app.post("/activity-photo-normalize")
async def activity_photo_normalize(
    laptop_number: str = Form(...),
    schedule_id: int | None = Form(None),
    schedule_code: str = Form(""),
    schedule_time: str = Form(""),
    schedule_time_range: str = Form(""),
    schedule_source: str = Form("api"),
    activities_json: str = Form(""),
    schedule_date: str = Form(""),
    schedule_content: str = Form(""),
    photographer: str = Form(""),
    normalize_mode: str = Form("exif"),
    source_folder: str = Form(""),
    output_folder: str = Form(""),
    backup_folder: str = Form(""),
    files: list[UploadFile] = File(...),
):
    if not files:
        raise HTTPException(status_code=400, detail="請先選擇活動照片。")

    for folder in (source_folder, output_folder, backup_folder):
        text = str(folder or "").strip()
        if text:
            with contextlib.suppress(Exception):
                Path(text).mkdir(parents=True, exist_ok=True)

    try:
        return await normalize_activity_photo_files(
            files=files,
            laptop_number=laptop_number,
            photographer=photographer,
            normalize_mode=normalize_mode or "exif",
            schedule_id=schedule_id,
            schedule_code=schedule_code,
            schedule_time=schedule_time,
            schedule_time_range=schedule_time_range,
            schedule_source=schedule_source,
            activities_json=activities_json,
            schedule_date=schedule_date,
            schedule_content=schedule_content,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"活動照片匯入失敗：{str(exc)}") from exc


@app.post("/admin/open-log")
async def admin_open_log(payload: OpenLogPayload):
    mapping = {
        "normalize": WINDOWS_BATCH_NORMALIZE_LOG_PATH,
        "feature-build": FEATURE_BUILD_LOG_PATH,
        "activity-photo-import": ACTIVITY_PHOTO_IMPORT_LOG_PATH,
    }
    target = mapping.get(payload.log_name)
    if target is None:
        raise HTTPException(status_code=400, detail="不支援的 log 類型。")
    if not target.exists():
        raise HTTPException(status_code=404, detail=f"找不到 log 檔案：{target}")
    subprocess.Popen(["notepad.exe", str(target)])
    return {"ok": True, "log_path": str(target)}


@app.post("/admin/open-folder")
async def admin_open_folder(payload: OpenFolderPayload):
    folder = Path(str(payload.folder_path or "").strip())
    if not str(folder):
        raise HTTPException(status_code=400, detail="請提供資料夾路徑")
    with contextlib.suppress(Exception):
        folder.mkdir(parents=True, exist_ok=True)
    if not folder.exists() or not folder.is_dir():
        raise HTTPException(status_code=404, detail=f"資料夾不存在：{folder}")
    subprocess.Popen(["explorer.exe", str(folder)])
    return {"ok": True, "folder_path": str(folder)}


@app.post("/admin/migrate-legacy-logs")
async def admin_migrate_legacy_logs():
    ensure_noob_log_root()
    LEGACY_MIGRATED_LOG_ROOT.mkdir(parents=True, exist_ok=True)

    moved_files = []
    if LEGACY_LOG_ROOT.exists():
        for src in sorted(LEGACY_LOG_ROOT.glob("*.log"), key=lambda p: p.name.lower()):
            target = LEGACY_MIGRATED_LOG_ROOT / src.name
            if target.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                target = LEGACY_MIGRATED_LOG_ROOT / f"{src.stem}_{timestamp}{src.suffix}"
            with contextlib.suppress(Exception):
                shutil.move(str(src), str(target))
                moved_files.append(str(target))

    return {
        "ok": True,
        "legacy_log_root": str(LEGACY_LOG_ROOT),
        "migrated_root": str(LEGACY_MIGRATED_LOG_ROOT),
        "moved_count": len(moved_files),
        "moved_files": moved_files[:200],
    }


@app.post("/admin/upload-excel")
async def admin_upload_excel(file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="請選擇 Excel 或 CSV 檔案。")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".xlsx", ".xls", ".xlsm", ".xltx", ".xltm", ".csv"}:
        raise HTTPException(status_code=400, detail="僅支援 Excel 或 CSV 檔案。")

    job_dir = create_batch_job_dir(BATCH_UPLOAD_ROOT, "excel")
    target_path = job_dir / Path(file.filename).name
    save_upload_file(file, target_path)
    sheet_names = list_tabular_sheets(target_path)
    selected_sheet = "" if suffix == ".csv" else sheet_names[0]
    columns = load_excel_columns(target_path, sheet_name=selected_sheet)

    return {
        "filename": Path(file.filename).name,
        "server_path": str(target_path),
        "host_path": str(target_path),
        "sheet_names": sheet_names,
        "selected_sheet": selected_sheet or "CSV",
        "columns": columns,
    }


@app.get("/admin/excel-columns")
async def admin_excel_columns(
    excel_path: str = Query(..., min_length=1),
    sheet_name: str = Query("", description="工作表名稱"),
):
    sheet_names = list_tabular_sheets(excel_path)
    normalized_sheet = ""
    if sheet_names != ["CSV"]:
        normalized_sheet = sheet_name or sheet_names[0]
        if normalized_sheet not in sheet_names:
            raise HTTPException(status_code=400, detail=f"找不到工作表：{normalized_sheet}")
    columns = load_excel_columns(excel_path, sheet_name=normalized_sheet)
    return {
        "sheet_names": sheet_names,
        "selected_sheet": normalized_sheet or "CSV",
        "columns": columns,
    }


@app.post("/admin/upload-source-folder")
async def admin_upload_source_folder(files: list[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="請選擇來源圖檔資料夾。")

    source_dir = create_batch_job_dir(BATCH_UPLOAD_ROOT, "source")
    suggested_destination_dir = str(BATCH_OUTPUT_ROOT / f"{source_dir.name}_normalized")
    saved_count = 0

    for upload in files:
        if not upload.filename:
            continue
        target_path = source_dir / Path(upload.filename).name
        save_upload_file(upload, target_path)
        saved_count += 1

    if saved_count == 0:
        raise HTTPException(status_code=400, detail="來源圖檔資料夾內沒有可上傳的檔案。")

    return {
        "server_path": str(source_dir),
        "host_path": str(source_dir),
        "file_count": saved_count,
        "suggested_destination_dir": suggested_destination_dir,
        "suggested_destination_host_dir": suggested_destination_dir,
    }


@app.post("/admin/batch-normalize")
async def admin_batch_normalize(payload: BatchNormalizePayload):
    try:
        return normalize_headshot_batch(payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except PermissionError as exc:
        target_path = str(exc.filename) if getattr(exc, "filename", None) else payload.destination_dir
        if str(target_path).lower().startswith(r"c:\feature_src"):
            detail = r"C:\feature_src 無寫入權限，請改用其他目的圖檔資料夾或調整資料夾權限。"
        else:
            detail = f"目的圖檔資料夾無寫入權限：{target_path}"
        raise HTTPException(status_code=400, detail=detail) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"活動照片匯入失敗：{str(exc)}") from exc
@app.post("/activity-photo-normalize-folder")
async def activity_photo_normalize_folder_api(
    laptop_number: str = Form(...),
    schedule_id: int | None = Form(None),
    schedule_code: str = Form(""),
    schedule_time: str = Form(""),
    schedule_time_range: str = Form(""),
    schedule_source: str = Form("api"),
    activities_json: str = Form(""),
    schedule_date: str = Form(""),
    schedule_content: str = Form(""),
    photographer: str = Form(""),
    normalize_mode: str = Form("exif"),
    source_folder: str = Form(""),
    output_folder: str = Form(""),
):
    text = str(source_folder or "").strip()
    output_text = str(output_folder or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="請先填入來源圖檔資料夾。")
    if text.replace("/", "\\").lower().startswith(r"c:\uploadsource") or output_text.replace("/", "\\").lower().startswith(r"c:\uploadsource"):
        raise HTTPException(status_code=400, detail="請改用 C:\\activity\\ingest 目錄，不可混用舊路徑 C:\\uploadsource。")
    with contextlib.suppress(Exception):
        Path(text).mkdir(parents=True, exist_ok=True)
    if output_text:
        with contextlib.suppress(Exception):
            Path(output_text).mkdir(parents=True, exist_ok=True)
    try:
        return await normalize_activity_photo_folder(
            source_folder=text,
            output_folder=output_text or r"C:\activity\ingest\normalized_success",
            laptop_number=laptop_number,
            schedule_id=schedule_id,
            schedule_code=schedule_code,
            schedule_time=schedule_time,
            schedule_time_range=schedule_time_range,
            schedule_source=schedule_source,
            activities_json=activities_json,
            schedule_date=schedule_date,
            schedule_content=schedule_content,
            photographer=photographer,
            normalize_mode=normalize_mode or "exif",
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"活動照片匯入失敗：{str(exc)}") from exc


@app.post("/activity-photo-normalize/start")
async def activity_photo_normalize_start_proxy(
    laptop_number: str = Form(""),
    schedule_id: str = Form(""),
    schedule_code: str = Form(""),
    schedule_time: str = Form(""),
    schedule_time_range: str = Form(""),
    schedule_source: str = Form("api"),
    activities_json: str = Form(""),
    schedule_date: str = Form(""),
    schedule_content: str = Form(""),
    photographer: str = Form(""),
    normalize_mode: str = Form("exif"),
    source_folder: str = Form(""),
    output_folder: str = Form(""),
):
    try:
        schedule_id_value = str(schedule_id or "").strip()
        if schedule_id_value.lower() in {"none", "null"}:
            schedule_id_value = ""
        payload = {
            "laptop_number": laptop_number,
            "schedule_code": schedule_code,
            "schedule_time": schedule_time,
            "schedule_time_range": schedule_time_range,
            "schedule_source": schedule_source,
            "activities_json": activities_json,
            "schedule_date": schedule_date,
            "schedule_content": schedule_content,
            "photographer": photographer,
            "normalize_mode": normalize_mode,
            "source_folder": source_folder,
            "output_folder": output_folder,
        }
        if schedule_id_value and schedule_id_value.isdigit():
            payload["schedule_id"] = schedule_id_value
        encoded = urllib.parse.urlencode(payload).encode("utf-8")
        request = urllib.request.Request(
            "http://127.0.0.1:8000/activity-photo-normalize/start",
            data=encoded,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            method="POST",
        )
        with urllib.request.urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_proxy_error_message(exc, "活動照片正規化任務啟動")) from exc


@app.get("/activity-photo-normalize/jobs/{job_id}")
async def activity_photo_normalize_job_status_proxy(job_id: str):
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:8000/activity-photo-normalize/jobs/{job_id}", timeout=20) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_proxy_error_message(exc, "查詢正規化任務狀態")) from exc


@app.get("/activity-photo-normalize/jobs/{job_id}/logs")
async def activity_photo_normalize_job_logs_proxy(job_id: str, offset: int = Query(0)):
    try:
        safe_offset = max(0, int(offset or 0))
        url = f"http://127.0.0.1:8000/activity-photo-normalize/jobs/{job_id}/logs?offset={safe_offset}"
        with urllib.request.urlopen(url, timeout=20) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_proxy_error_message(exc, "查詢正規化任務 Log")) from exc


@app.get("/activity-photo-normalize/jobs-recent")
async def activity_photo_normalize_jobs_recent_proxy(limit: int = Query(30)):
    try:
        safe_limit = max(1, min(int(limit or 30), 200))
        with urllib.request.urlopen(f"http://127.0.0.1:8000/activity-photo-normalize/jobs-recent?limit={safe_limit}", timeout=20) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_proxy_error_message(exc, "讀取最近正規化任務")) from exc


@app.post("/activity-photo-import")
async def activity_photo_import_api(
    laptop_number: str = Form(...),
    schedule_id: int | None = Form(None),
    photographer: str = Form(""),
    enable_pyiqa: bool = Form(False),
    normalize_mode: str = Form("schedule"),
    source_folder: str = Form(""),
    output_folder: str = Form(""),
    backup_folder: str = Form(""),
):
    _ = source_folder, output_folder, backup_folder
    try:
        return await import_activity_photos_from_normalized_folder(
            laptop_number=laptop_number,
            schedule_id=schedule_id,
            photographer=photographer,
            enable_pyiqa=enable_pyiqa,
            normalize_mode=normalize_mode,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"活動照片匯入失敗：{str(exc)}") from exc
@app.post("/activity-photo-import-retry-failed")
@app.post("/activity-photo-import-retry-failed-proxy")
async def activity_photo_import_retry_failed_api(limit: int = Form(100)):
    try:
        safe_limit = max(1, min(int(limit or 100), 1000))
        form = {"limit": str(safe_limit)}
        data = urllib.parse.urlencode(form).encode("utf-8")
        req = urllib.request.Request(
            "http://127.0.0.1:8000/activity-photo-import-retry-failed",
            data=data,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )
        with urllib.request.urlopen(req, timeout=600) as response:
            body = response.read().decode("utf-8", "replace")
            return json.loads(body)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_proxy_error_message(exc, "活動照片辨識補跑")) from exc
@app.get("/activity-photo-import/jobs/{job_id}")
async def activity_photo_import_job_status_proxy(job_id: str):
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:8000/activity-photo-import/jobs/{job_id}", timeout=20) as response:
            body = response.read().decode("utf-8", "replace")
            return json.loads(body)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_proxy_error_message(exc, "查詢匯入任務狀態")) from exc


@app.get("/activity-photo-import/jobs/{job_id}/logs")
async def activity_photo_import_job_logs_proxy(job_id: str, offset: int = Query(0)):
    try:
        url = f"http://127.0.0.1:8000/activity-photo-import/jobs/{job_id}/logs?offset={max(0, int(offset or 0))}"
        with urllib.request.urlopen(url, timeout=20) as response:
            body = response.read().decode("utf-8", "replace")
            return json.loads(body)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_proxy_error_message(exc, "查詢匯入任務 Log")) from exc


@app.get("/activity-photo-import/jobs/{job_id}/items")
async def activity_photo_import_job_items_proxy(job_id: str):
    try:
        with urllib.request.urlopen(f"http://127.0.0.1:8000/activity-photo-import/jobs/{job_id}/items", timeout=20) as response:
            body = response.read().decode("utf-8", "replace")
            return json.loads(body)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_proxy_error_message(exc, "查詢匯入任務明細")) from exc


@app.post("/activity-photo-import-proxy")
async def activity_photo_import_proxy_api(
    laptop_number: str = Form(...),
    schedule_id: int | None = Form(None),
    photographer: str = Form(""),
    enable_pyiqa: bool = Form(False),
    normalize_mode: str = Form("schedule"),
    source_folder: str = Form(""),
    output_folder: str = Form(""),
    backup_folder: str = Form(""),
):
    try:
        # 僅啟動 server 端 job，不在 8010 先做整批 bridge 複製，避免大批次 timeout。
        resolved_source_folder = (source_folder or r"C:\activity\ingest\normalized_success").strip()
        if not resolved_source_folder:
            resolved_source_folder = r"C:\activity\ingest\normalized_success"
        if resolved_source_folder.replace("/", "\\").lower().startswith(r"c:\uploadsource"):
            raise HTTPException(status_code=400, detail="請改用 C:\\activity\\ingest 目錄，不可混用舊路徑 C:\\uploadsource。")

        form = {
            "laptop_number": laptop_number,
            "photographer": photographer or "",
            "enable_pyiqa": "true" if enable_pyiqa else "false",
            "normalize_mode": normalize_mode or "schedule",
            "source_folder": resolved_source_folder,
            "output_folder": output_folder or "",
            "backup_folder": backup_folder or "",
        }
        if schedule_id is not None:
            form["schedule_id"] = str(schedule_id)

        data = urllib.parse.urlencode(form).encode("utf-8")
        req = urllib.request.Request(
            "http://127.0.0.1:8000/activity-photo-import/start",
            data=data,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )
        with urllib.request.urlopen(req, timeout=60) as response:
            body = response.read().decode("utf-8", "replace")
            payload = json.loads(body)

        if not isinstance(payload, dict):
            raise HTTPException(status_code=500, detail="活動照片匯入啟動失敗：回傳格式異常")

        payload["windows_source_folder"] = resolved_source_folder
        payload["windows_mode"] = "start_and_poll"
        return payload
    except Exception as exc:
        raise HTTPException(status_code=500, detail=_proxy_error_message(exc, "活動照片匯入")) from exc

